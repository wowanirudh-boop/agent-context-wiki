"""Local document processor — runs extraction without S3 or Postgres.

Processes files directly from the workspace filesystem and updates SQLite.
Respects PDF_BACKEND config and optional Mistral/LibreOffice backends.
"""

import asyncio
import json
import logging
import shutil
import subprocess
import tempfile
import uuid
from pathlib import Path

import aiosqlite

from config import settings
from domain.watcher import mark_written
from infra.db.sqlite import SQLiteDocumentRepository
from services.extracted_assets import build_pdf_image_assets

logger = logging.getLogger(__name__)

PDF_TYPES = frozenset({"pdf"})
SPREADSHEET_TYPES = frozenset({"xlsx", "xls"})
IMAGE_TYPES = frozenset({"png", "jpg", "jpeg", "webp", "gif"})
OFFICE_TYPES = frozenset({"pptx", "ppt", "docx", "doc"})

from services.pdf_extract import extract_pdf


async def process_document(db: aiosqlite.Connection, doc_id: str, workspace: Path) -> None:
    """Process a pending document: extract text, chunk, update index."""
    cursor = await db.execute(
        "SELECT id, filename, file_type, relative_path, status FROM documents WHERE id = ?",
        (doc_id,),
    )
    row = await cursor.fetchone()
    if not row:
        logger.warning("Document %s not found", doc_id[:8])
        return

    cols = [d[0] for d in cursor.description]
    doc = dict(zip(cols, row))

    if doc["status"] not in ("pending", "processing"):
        return

    file_type = doc["file_type"] or ""
    file_path = workspace / doc["relative_path"]

    if not file_path.is_file():
        await db.execute(
            "UPDATE documents SET status = 'failed', error_message = 'File not found', "
            "updated_at = datetime('now') WHERE id = ?",
            (doc_id,),
        )
        await db.commit()
        return

    await db.execute(
        "UPDATE documents SET status = 'processing', updated_at = datetime('now') WHERE id = ?",
        (doc_id,),
    )
    await db.commit()

    try:
        if file_type in PDF_TYPES:
            await _process_pdf(db, doc_id, file_path, workspace)
        elif file_type in OFFICE_TYPES:
            await _process_office(db, doc_id, file_path, workspace)
        elif file_type in SPREADSHEET_TYPES:
            await _process_spreadsheet(db, doc_id, file_path)
        elif file_type in IMAGE_TYPES:
            await _process_image(db, doc_id)
        elif file_type in ("html", "htm"):
            await _process_html(db, doc_id, file_path)
        else:
            await db.execute(
                "UPDATE documents SET status = 'ready', updated_at = datetime('now') WHERE id = ?",
                (doc_id,),
            )
            await db.commit()

        logger.info("Processed %s: %s", doc["filename"], file_type)

    except Exception as e:
        error_msg = str(e)[:500]
        await db.execute(
            "UPDATE documents SET status = 'failed', error_message = ?, "
            "updated_at = datetime('now') WHERE id = ?",
            (error_msg, doc_id),
        )
        await db.commit()
        logger.error("Failed to process %s: %s", doc["filename"], e)


async def _store_chunks(db: aiosqlite.Connection, doc_id: str, chunks: list) -> None:
    """Store chunks into SQLite, replacing any existing ones."""
    await db.execute("DELETE FROM document_chunks WHERE document_id = ?", (doc_id,))
    for c in chunks:
        await db.execute(
            "INSERT INTO document_chunks (id, document_id, chunk_index, content, source_content, page, "
            "start_char, token_count, header_breadcrumb) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), doc_id, c.index, c.content, c.content, c.page,
             c.start_char, c.token_count, c.header_breadcrumb),
        )


# ── PDF extraction ────────────────────────────────────────────────────────

async def _save_local_images(
    db: aiosqlite.Connection, doc_id: str, workspace: Path,
    pages_with_images: list[tuple[int, str, list[dict]]],
) -> dict[int, dict]:
    """Save extracted images as hidden sibling assets and return page metadata."""
    repo = SQLiteDocumentRepository(db)
    doc = await repo.get(doc_id)
    if not doc:
        return {}

    assets, page_elements = build_pdf_image_assets(
        doc_id,
        doc["filename"],
        doc["path"],
        pages_with_images,
    )
    if not assets:
        return {}

    await db.execute(
        "DELETE FROM documents WHERE source_kind = 'asset' AND metadata LIKE ?",
        (f'%"parent_document_id": "{doc_id}"%',),
    )
    await db.commit()

    asset_metadata = []
    for asset in assets:
        relative_asset = (asset.path.rstrip("/") + "/" + asset.filename).lstrip("/")
        local_asset = workspace / relative_asset
        local_asset.parent.mkdir(parents=True, exist_ok=True)
        mark_written(str(local_asset))
        local_asset.write_bytes(asset.data)
        await repo.create_asset(
            asset.document_id,
            doc["user_id"],
            asset.filename,
            asset.path,
            asset.filename,
            asset.file_type,
            len(asset.data),
            asset.metadata(),
        )
        asset_metadata.append(asset.metadata())

    await repo.set_metadata_field(doc_id, "assets", asset_metadata)
    return page_elements


async def _store_page_contents(
    db: aiosqlite.Connection, doc_id: str,
    page_contents: list[tuple[int, str]], parser: str,
    page_elements: dict[int, dict] | None = None,
) -> None:
    """Store extracted pages, chunks, and update document status."""
    num_pages = len(page_contents)

    await db.execute("DELETE FROM document_pages WHERE document_id = ?", (doc_id,))
    for page_num, content in page_contents:
        elements = (page_elements or {}).get(page_num)
        await db.execute(
            "INSERT INTO document_pages (id, document_id, page, content, elements) VALUES (?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), doc_id, page_num, content,
             json.dumps(elements) if elements else None),
        )

    full_content = "\n\n---\n\n".join(md for _, md in page_contents)

    from services.chunker import chunk_pages
    chunks = chunk_pages(page_contents)
    await _store_chunks(db, doc_id, chunks)

    await db.execute(
        "UPDATE documents SET status = 'ready', content = ?, page_count = ?, "
        "parser = ?, updated_at = datetime('now') WHERE id = ?",
        (full_content, num_pages, parser, doc_id),
    )
    await db.commit()
    from core.ingest import seed_existing_document_chunks
    await seed_existing_document_chunks(db, doc_id, workspace=None, content=full_content)


async def _process_pdf(db: aiosqlite.Connection, doc_id: str, file_path: Path, workspace: Path) -> None:
    """Extract PDF text. Uses opendataloader by default, Mistral if configured."""
    if settings.PDF_BACKEND == "mistral" and settings.MISTRAL_API_KEY:
        await _process_pdf_mistral(db, doc_id, file_path, workspace)
    else:
        pages_with_images = await asyncio.to_thread(extract_pdf, str(file_path))
        page_elements = await _save_local_images(db, doc_id, workspace, pages_with_images)
        page_contents = [(num, md) for num, md, _ in pages_with_images]
        await _store_page_contents(db, doc_id, page_contents, "opendataloader", page_elements)


# ── Office processing ─────────────────────────────────────────────────────

async def _process_office(db: aiosqlite.Connection, doc_id: str, file_path: Path, workspace: Path) -> None:
    """Convert Office docs to PDF via local LibreOffice, then extract text."""
    lo = shutil.which("libreoffice") or shutil.which("soffice")
    if not lo:
        await db.execute(
            "UPDATE documents SET status = 'failed', "
            "error_message = 'LibreOffice not installed. Install it to process Office files.', "
            "updated_at = datetime('now') WHERE id = ?",
            (doc_id,),
        )
        await db.commit()
        return

    with tempfile.TemporaryDirectory() as tmpdir:
        result = await asyncio.to_thread(
            subprocess.run,
            [lo, "--headless", "--convert-to", "pdf", "--outdir", tmpdir, str(file_path)],
            capture_output=True, timeout=120,
        )
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice conversion failed: {result.stderr.decode()[:300]}")

        pdf_files = list(Path(tmpdir).glob("*.pdf"))
        if not pdf_files:
            raise RuntimeError("LibreOffice produced no PDF output")

        converted_pdf = pdf_files[0]

        # Store converted PDF in cache for the viewer
        cache_dir = workspace / ".llmwiki" / "cache" / "local" / doc_id
        cache_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(converted_pdf, cache_dir / "converted.pdf")

        pages_with_images = await asyncio.to_thread(extract_pdf, str(converted_pdf))
        page_elements = await _save_local_images(db, doc_id, workspace, pages_with_images)
        page_contents = [(num, md) for num, md, _ in pages_with_images]
        await _store_page_contents(db, doc_id, page_contents, "libreoffice+opendataloader", page_elements)


# ── Mistral OCR ───────────────────────────────────────────────────────────

async def _process_pdf_mistral(db: aiosqlite.Connection, doc_id: str, file_path: Path, workspace: Path) -> None:
    """Extract PDF via Mistral OCR API (better tables/layout, requires API key)."""
    import httpx
    import base64

    pdf_bytes = file_path.read_bytes()
    pdf_b64 = base64.b64encode(pdf_bytes).decode()

    async with httpx.AsyncClient(timeout=120) as client:
        resp = await client.post(
            "https://api.mistral.ai/v1/ocr",
            headers={"Authorization": f"Bearer {settings.MISTRAL_API_KEY}"},
            json={
                "model": "mistral-ocr-latest",
                "document": {"type": "document_url", "document_url": f"data:application/pdf;base64,{pdf_b64}"},
            },
        )
        resp.raise_for_status()
        result = resp.json()

    pages = result.get("pages", [])
    page_contents = [(i + 1, p.get("markdown", "")) for i, p in enumerate(pages)]
    await _store_page_contents(db, doc_id, page_contents, "mistral")


# ── Spreadsheet processing ────────────────────────────────────────────────

async def _process_spreadsheet(db: aiosqlite.Connection, doc_id: str, file_path: Path) -> None:
    """Extract spreadsheet data via openpyxl. Stores pages AND chunks for search."""
    from openpyxl import load_workbook

    wb = await asyncio.to_thread(load_workbook, str(file_path), read_only=True, data_only=True)

    await db.execute("DELETE FROM document_pages WHERE document_id = ?", (doc_id,))

    all_content = []
    page_contents = []
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(" | ".join(str(c) if c is not None else "" for c in row))
        content = "\n".join(rows)
        elements = json.dumps({"sheet_name": sheet_name})

        await db.execute(
            "INSERT INTO document_pages (id, document_id, page, content, elements) "
            "VALUES (?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), doc_id, i, content, elements),
        )
        all_content.append(f"## {sheet_name}\n\n{content}")
        page_contents.append((i, content))

    num_sheets = len(wb.sheetnames)
    wb.close()
    full_content = "\n\n".join(all_content)

    from services.chunker import chunk_pages
    chunks = chunk_pages(page_contents)
    await _store_chunks(db, doc_id, chunks)

    await db.execute(
        "UPDATE documents SET status = 'ready', content = ?, page_count = ?, "
        "parser = 'openpyxl', updated_at = datetime('now') WHERE id = ?",
        (full_content, num_sheets, doc_id),
    )
    await db.commit()
    from core.ingest import seed_existing_document_chunks
    await seed_existing_document_chunks(db, doc_id, content=full_content)


# ── Image / HTML processing ──────────────────────────────────────────────

async def _process_image(db: aiosqlite.Connection, doc_id: str) -> None:
    """Images are stored as-is — just mark ready."""
    await db.execute(
        "UPDATE documents SET status = 'ready', page_count = 1, "
        "parser = 'native', updated_at = datetime('now') WHERE id = ?",
        (doc_id,),
    )
    await db.commit()


async def _process_html(db: aiosqlite.Connection, doc_id: str, file_path: Path) -> None:
    """Extract HTML content via webmd parser."""
    raw_html = file_path.read_text(encoding="utf-8", errors="replace")

    try:
        from html_parser import Parser
        parser = Parser(raw_html, content_only=True)
        result = parser.parse()
        content = result.content
    except Exception:
        content = raw_html

    from services.chunker import chunk_text
    chunks = chunk_text(content)
    await _store_chunks(db, doc_id, chunks)

    await db.execute(
        "UPDATE documents SET status = 'ready', content = ?, page_count = 1, "
        "parser = 'webmd', updated_at = datetime('now') WHERE id = ?",
        (content, doc_id),
    )
    await db.commit()
    from core.ingest import seed_existing_document_chunks
    await seed_existing_document_chunks(db, doc_id, content=content)
